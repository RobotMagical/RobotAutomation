import openpyxl


def update_result_to_inputexcel(infile, sheetname, datamap):
    infile = infile + ".xlsx"
    currFile = openpyxl.load_workbook(infile)
    currSheet = currFile[sheetname]
    row = 2
    pairlist = datamap.split(',')
    for i in range(len(pairlist)):
        column_name = "${"+(pairlist[i].split('='))[0]+"}"
        column_value = (pairlist[i].split('='))[1]
        specificCellLetter = (find_specific_cell(currSheet, column_name))
        letter = get_column_letter(specificCellLetter)
        # get_all_values_by_cell_letter(letter)
        append_cell_value(infile, currFile, currSheet, letter, row, column_value)


def append_result_to_outputexcel(infile, sheetname, datamap):
    infile = infile + ".xlsx"
    currFile = openpyxl.load_workbook(infile)
    currSheet = currFile[sheetname]
    row = currSheet.max_row + 1
    pairlist = datamap.split(',')
    for i in range(len(pairlist)):
        column_name = (pairlist[i].split('='))[0]
        column_value = (pairlist[i].split('='))[1]
        specificCellLetter = (find_specific_cell(currSheet, column_name))
        letter = get_column_letter(specificCellLetter)
        # get_all_values_by_cell_letter(letter)
        write_cell_value(infile, currFile, currSheet, letter, row, column_value)


def find_specific_cell(currentSheet, column_name):
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGHIJKL":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == column_name:
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name


def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    # print(letter)
    return letter


def get_all_values_by_cell_letter(currentSheet, letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            # print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))


def append_cell_value(infile, theFile, currentSheet, letter, rownum, value):
    # row = currentSheet.max_row + 1
    for column in letter:
        cell_name = "{}{}".format(column, rownum)
        currentSheet[cell_name].value = value
    theFile.save(infile)


def write_cell_value(infile, theFile, currentSheet, letter, rownum, value):
    # row = currentSheet.max_row + 1
    for column in letter:
        cell_name = "{}{}".format(column, rownum)
        currentSheet[cell_name].value = value
    theFile.save(infile)
